# -*- coding: utf-8 -*-
"""将 2-公式.csv 同步到 手动表.xlsx 的「2-公式表」；常数表 D_min 命名；士兵表验算列表头修正。"""
import csv
import os

from openpyxl import load_workbook


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base, "手动表.xlsx")
    csv_path = os.path.join(base, "2-公式.csv")
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path)

    ws = wb["2-公式表"]
    ws.delete_rows(1, ws.max_row)
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            out = (row + ["", "", "", ""])[:4]
            ws.append(out)

    ws3 = wb["3-常数表"]
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        name = row[0].value
        if name == "min_damage":
            row[0].value = "D_min"
            row[2].value = "减法公式伤害下限 max(D_min,基础伤害−防御)；与策划文档统一命名"

    ws4 = wb["4-标准士兵表"]
    for col in range(1, 9):
        cell = ws4.cell(row=1, column=col)
        v = cell.value
        if v is None:
            continue
        s = str(v)
        for old, new in (
            ("攻击同等阶士兵刀数", "攻击同等阶敌人刀数"),
            ("攻击上一阶士兵刀数", "攻击上一阶敌人刀数"),
            ("攻击下一阶士兵刀数", "攻击下一阶敌人刀数"),
        ):
            if old in s:
                cell.value = s.replace(old, new)

    wb.save(path)
    print("已更新:", path)


if __name__ == "__main__":
    main()
