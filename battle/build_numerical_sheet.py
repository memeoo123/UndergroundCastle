# -*- coding: utf-8 -*-
"""生成地下城堡数值设计表 Excel（基础属性说明 / 公式 / 推导数值 / 单位配置）"""
import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装: pip install openpyxl")
    raise

def style_header(ws):
    """表头样式"""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sheet_1_basic_attrs(wb):
    """分表1：基础属性说明"""
    ws = wb.create_sheet("1-基础属性说明", 0)
    headers = ["属性ID", "中文名", "说明", "单位", "取值范围/示例", "用于", "备注"]
    rows = [
        ["max_hp", "生命值", "单位存活时的生命上限", "点", "100~10000+", "战斗", "为0即死亡"],
        ["atk", "攻击力", "用于伤害计算的基础攻击", "点", "10~5000+", "战斗", "参与伤害公式"],
        ["def", "防御力", "减法中与攻击/基础伤害配对相减", "点", "0~3000+", "战斗", "主方案：减法"],
        ["spd", "速度", "ATB时间条增长速度；速度=10 时为 1 秒 1 刀，按兵种调此值即可调节奏", "点", "约 5~20+；基准10", "战斗", "见常数表 speed_baseline"],
        ["crit_rate", "暴击率", "暴击判定概率", "百分比", "0%~100%", "战斗", "可设上限"],
        ["crit_dmg", "暴击伤害", "暴击时的伤害倍率", "倍率", "1.5~3.0", "战斗", "如150%即1.5倍"],
        ["hit", "命中", "命中判定（可选）", "点或%", "0~100%", "战斗", "若不做闪避可省略"],
        ["dodge", "闪避", "闪避判定（可选）", "点或%", "0~100%", "战斗", "与命中配套"],
        ["--", "--", "--", "--", "--", "--", "--"],
        ["刀数(节奏)", "刀数", "击败一只怪所需攻击次数 = 怪物血量/单次伤害", "次", "设计目标约10", "战斗节奏", "用于反推血量、控制单场战斗时长"],
        ["--", "--", "--", "--", "--", "--", "--"],
        ["stamina", "体力/耐力", "地牢探险消耗（可选）", "点", "100", "地牢", "种田或休息恢复"],
        ["dungeon_party_max", "地牢队伍上限", "单次进入地牢可编入队伍的单位人数上限；超出则不可选入或不可开始探险", "人", "1~6（示例）", "地牢", "全局默认见常数表；关卡可覆盖；影响 ATB 同场单位数"],
        ["--", "--", "--", "--", "--", "--", "--"],
        ["growth_days", "成熟天数", "种田作物成熟所需天数", "天", "1~30", "种田", "可做表单独配置"],
        ["yield_count", "单次产量", "收获时产出数量", "个", "1~99", "种田", "按作物ID配置"],
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    style_header(ws)
    ws.freeze_panes = "A2"

def sheet_2_formulas(wb):
    """分表2：公式（减法主方案）"""
    ws = wb.create_sheet("2-公式", 1)
    headers = ["公式名称", "公式表达式（策划可调）", "参数说明", "备注"]
    rows = [
        ["--", "--", "--", "--"],
        ["【主方案】减法公式（约10阶封闭环境）", "--", "--", "--"],
        ["减法-单次有效伤害", "单次有效伤害 = max(D_min，基础伤害 − 守方防御)", "D_min：伤害下限（建议1）；防御为受击方防御", "技能倍率、克制可先=1做表"],
        ["减法-基础伤害（进攻方）", "基础伤害 = 攻击 × 技能倍率 × (1 ± 克制系数) × 上阵人数", "士兵侧含人数；敌人单体人数=1", "与 dungeon_party_max 校验"],
        ["减法-刀数", "刀数 = ⌈目标血量 / 单次有效伤害⌉", "单次有效伤害为减法结果", "先定同阶刀数再反推血量"],
        ["减法-反推血量", "目标血量 ≈ 目标刀数 × 单次有效伤害", "先定刀数与每刀伤害", "按减法反推即可"],
        ["减法-敌方打我方阵（可选）", "单次伤害=max(D_min，敌攻−我防)；刀数=⌈(我单兵血×人数)/单次伤害⌉", "总血池简化", "与士兵打敌对称"],
        ["减法-验算乘序（重要）", "「先差再乘」：(攻−防)×人数；「先乘再减」：攻×人数−防", "人数>1 时一般不等", "手动表与程序统一一种"],
        ["--", "--", "--", "--"],
        ["暴击判定", "若 rand() < 暴击率 则 伤害 × 暴击伤害", "rand()为0~1随机数", "可设暴击率上限如80%"],
        ["刀数（节奏指标）", "刀数 = 怪物血量 / 玩家单次有效伤害（向上取整）", "即击败一只怪所需攻击次数；用于控制战斗节奏", "设计目标：普通战约10刀，简单6~8刀，难12~15刀，Boss可更高"],
        ["反推血量", "怪物血量 = 目标刀数 × 期望单次伤害", "先定目标刀数与单次伤害，反推怪物血量", "便于按节奏做表"],
        ["ATB时间条增速", "每秒增加量 = 速度 × atb_speed_factor（实现时用 delta_time 累加，与 FPS 无关）", "速度=10 时 1 秒满条 = 1 秒 1 刀；按兵种调速度即调攻击节奏", "atb_speed_factor = atb_max / speed_baseline，见常数表"],
        ["1秒1刀（基准）", "速度=10 时每秒增加 atb_max → 1 秒行动一次；其余速度按比例缩放", "实现方式与数值表分离", "如速度 5≈2 秒 1 刀，速度 20≈0.5 秒 1 刀"],
        ["行动后时间条", "行动后 时间条 = 0（或 满值 - 行动消耗）", "若技能有「占条」则扣对应量", "由程序实现，此处仅说明"],
        ["地牢编队限制", "上阵人数 ≤ min(玩家选中人数, dungeon_party_max)；关卡可设 party_max_override", "dungeon_party_max 见常数表；与基础属性说明一致", "进入地牢前校验；影响 ATB 同场友方单位数"],
        ["--", "--", "--", "--"],
        ["治疗量", "治疗量 = 治疗基础值 × (1 + 治疗加成%)", "治疗基础值来自技能或属性", "可再受防御/抗性影响（若需要）"],
        ["护盾值", "护盾 = 吸收量，优先扣除护盾再扣血", "吸收量来自技能或属性", "多源护盾可叠加或取高由设计定"],
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 36
    style_header(ws)
    ws.freeze_panes = "A2"

def sheet_3_derived(wb):
    """分表3：推导的数值（示例：地牢层级 + 常数）"""
    ws = wb.create_sheet("3-推导数值", 2)
    # 3.1 常数表
    ws.append(["【常数表】策划可调参数"])
    ws.append(["参数名", "建议值", "说明"])
    ws.append(["atb_max", 10000, "ATB时间条满值（满即行动）"])
    ws.append(["speed_baseline", 10, "基准速度：速度=10 时 1 秒满条 = 1 秒 1 刀；按兵种调速度即可调节奏"])
    ws.append(["atb_speed_factor", 1000, "每秒增量 = 速度×此系数；atb_speed_factor = atb_max/speed_baseline"])
    ws.append(["target_hits_normal", 10, "普通战斗目标刀数（击败单怪所需攻击次数），用于反推血量与节奏"])
    ws.append(["dungeon_party_max", 4, "地牢探险编队人数上限（全局默认）；关卡/难度表可单独覆盖"])
    ws.append(["D_min", 1, "减法公式伤害下限 max(D_min,基础−防御)；避免伤害为0；封闭10阶可调"])
    ws.append(["battle_formula", "减法", "主战斗伤害公式：减法（攻防差），不用减伤率公式"])
    ws.append([])
    # 3.2 地牢层级示例（阶梯感：10层一档，量级拉开；刀数控制在约10刀）
    ws.append(["【地牢层级示例】量级阶梯（按目标刀数≈10 反推血量）"])
    ws.append(["层数", "档位", "怪物攻击", "怪物防御", "怪物血量", "推荐玩家攻击", "推荐玩家防御", "期望单次伤害(约)", "期望刀数", "备注"])
    tier_rows = [
        [5, "1档", 40, 20, 350, 60, 25, 35, 10, "入门"],
        [10, "1档", 80, 40, 700, 120, 50, 70, 10, "1档顶"],
        [15, "2档", 180, 90, 1400, 250, 120, 140, 10, "2档"],
        [20, "2档", 500, 250, 3400, 600, 200, 340, 10, "2档顶"],
        [25, "3档", 1200, 600, 8000, 1400, 500, 800, 10, "3档"],
        [30, "3档", 2000, 1000, 13300, 2400, 800, 1330, 10, "3档顶"],
    ]
    for r in tier_rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    for c in range(3, 11):
        ws.column_dimensions[get_column_letter(c)].width = 14
    # 只给第一个数据块加表头样式（常数表）
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.border = border
        cell.fill = fill
        cell.font = font
    for c in range(1, 11):
        cell = ws.cell(row=13, column=c)
        cell.border = border
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A14"


def _is_empty_row(r):
    return not r or all((c or "").strip() == "" for c in r)


def _coerce_unit_row(row, expected_len=11):
    """CSV 读入后为字符串，转数值列便于 Excel 里计算/筛选。"""
    out = list(row)
    while len(out) < expected_len:
        out.append("")
    # 生命、攻击、防御、速度
    for idx in range(4, 8):
        s = (out[idx] or "").strip()
        if s != "":
            try:
                out[idx] = int(float(s))
            except ValueError:
                pass
    for idx in (8, 9):
        s = (out[idx] or "").strip()
        if s != "":
            try:
                out[idx] = float(s)
            except ValueError:
                pass
    return out


def load_sheet4_from_csv(csv_path):
    """
    解析 4-敌人与士兵配置.csv：
    - 表头行第一列为「敌人ID」「兵种ID」
    - 敌人数据：敌人ID 表头下一行起，遇空行结束
    - 我方数据：兵种ID 表头下一行起，遇空行结束
    """
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    enemy_header_idx = None
    soldier_header_idx = None
    for i, r in enumerate(rows):
        if r and r[0].strip() == "敌人ID":
            enemy_header_idx = i
        if r and r[0].strip() == "兵种ID":
            soldier_header_idx = i

    if enemy_header_idx is None or soldier_header_idx is None:
        raise ValueError(
            "4-敌人与士兵配置.csv 中未找到表头「敌人ID」或「兵种ID」，请检查文件。"
        )
    if soldier_header_idx <= enemy_header_idx:
        raise ValueError("表头顺序错误：应先「敌人ID」段，后「兵种ID」段。")

    enemy_title = "【敌人配置表】"
    for k in range(enemy_header_idx - 1, -1, -1):
        r = rows[k]
        if _is_empty_row(r):
            continue
        if r[0].strip().startswith("【"):
            enemy_title = r[0].strip()
            break

    soldier_title = "【我方士兵配置表】"
    for k in range(soldier_header_idx - 1, enemy_header_idx, -1):
        r = rows[k]
        if _is_empty_row(r):
            continue
        if r[0].strip().startswith("【"):
            soldier_title = r[0].strip()
            break

    enemy_headers = rows[enemy_header_idx]
    enemy_rows_raw = []
    for j in range(enemy_header_idx + 1, soldier_header_idx):
        r = rows[j]
        if _is_empty_row(r):
            break
        if r[0].strip().startswith("【"):
            continue  # 若省略敌人与我方之间的空行，跳过小节标题行
        if r[0].strip() == "兵种ID":
            break
        enemy_rows_raw.append(r)

    soldier_headers = rows[soldier_header_idx]
    soldier_rows_raw = []
    for j in range(soldier_header_idx + 1, len(rows)):
        r = rows[j]
        if _is_empty_row(r):
            break
        soldier_rows_raw.append(r)

    enemy_rows = [_coerce_unit_row(r) for r in enemy_rows_raw]
    soldier_rows = [_coerce_unit_row(r) for r in soldier_rows_raw]

    return {
        "enemy_title": enemy_title,
        "enemy_headers": enemy_headers,
        "enemy_rows": enemy_rows,
        "soldier_title": soldier_title,
        "soldier_headers": soldier_headers,
        "soldier_rows": soldier_rows,
    }


def _apply_unit_block_style(ws, header_row, ncols, border, fill, font_header):
    """表头行（第 2 行）样式。"""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.border = border
        cell.fill = fill
        cell.font = font_header


def sheets_units_from_csv(wb, base_dir):
    """同一工作簿内两个页签：「4-敌人配置」「4-我方士兵配置」。数据源仍为单一 4-敌人与士兵配置.csv，未拆 CSV。"""
    csv_path = os.path.join(base_dir, "4-敌人与士兵配置.csv")
    if not os.path.isfile(csv_path):
        raise FileNotFoundError("缺少文件: " + csv_path)
    cfg = load_sheet4_from_csv(csv_path)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF")

    # ----- Sheet 4：敌人 -----
    ws_e = wb.create_sheet("4-敌人配置", 3)
    ws_e.append([cfg["enemy_title"]])
    ws_e.append(cfg["enemy_headers"])
    for r in cfg["enemy_rows"]:
        ws_e.append(r)
    ncols_e = max(len(cfg["enemy_headers"]), 11)
    _apply_unit_block_style(ws_e, 2, ncols_e, border, fill, font_header)
    for col in range(1, ncols_e + 1):
        ws_e.column_dimensions[get_column_letter(col)].width = 12
    ws_e.freeze_panes = "A3"

    # ----- 页签 2：我方（与敌人同属「第 4 部分」，同一 xlsx 内两个 Sheet）-----
    ws_s = wb.create_sheet("4-我方士兵配置", 4)
    ws_s.append([cfg["soldier_title"]])
    ws_s.append(cfg["soldier_headers"])
    for r in cfg["soldier_rows"]:
        ws_s.append(r)
    ncols_s = max(len(cfg["soldier_headers"]), 11)
    _apply_unit_block_style(ws_s, 2, ncols_s, border, fill, font_header)
    for col in range(1, ncols_s + 1):
        ws_s.column_dimensions[get_column_letter(col)].width = 12
    ws_s.freeze_panes = "A3"

def main():
    base = os.path.dirname(os.path.abspath(__file__))
    os.chdir(base)  # 确保输出到脚本所在目录
    path = os.path.join(base, "地下城堡数值表.xlsx")
    wb = Workbook()
    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    sheet_1_basic_attrs(wb)
    sheet_2_formulas(wb)
    sheet_3_derived(wb)
    sheets_units_from_csv(wb, base)
    wb.save(path)
    print("已生成:", path)

if __name__ == "__main__":
    main()
