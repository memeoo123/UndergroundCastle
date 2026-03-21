# -*- coding: utf-8 -*-
"""从 手动表.xlsx 导出「4-标准士兵表」「5-标准敌人表」为 JSON（取计算值，非公式文本）。"""
import json
import os

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装: pip install openpyxl")


def _cell_value(v):
    if v is None or v == "":
        return None
    return v


def sheet_to_list(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = []
    for h in rows[0]:
        s = "" if h is None else str(h).strip()
        headers.append(s)
    out = []
    for row in rows[1:]:
        if row is None or all((c is None or c == "") for c in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            d[h] = _cell_value(v)
        out.append(d)
    return out


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    xlsx = os.path.join(base, "手动表.xlsx")
    if not os.path.isfile(xlsx):
        raise FileNotFoundError(xlsx)

    wb = load_workbook(xlsx, data_only=True)

    for name, out_name in (
        ("4-标准士兵表", "手动表-4-标准士兵表.json"),
        ("5-标准敌人表", "手动表-5-标准敌人表.json"),
    ):
        if name not in wb.sheetnames:
            raise KeyError(f"工作簿中未找到: {name}，现有: {wb.sheetnames}")
        data = sheet_to_list(wb[name])
        path = os.path.join(base, out_name)
        payload = {
            "source": "手动表.xlsx",
            "sheet": name,
            "rowCount": len(data),
            "rows": data,
        }
        with open(path, "w", encoding="utf-8-sig", newline="\n") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print("已写入:", path)


if __name__ == "__main__":
    main()
